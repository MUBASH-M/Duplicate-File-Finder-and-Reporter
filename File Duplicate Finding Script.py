import os
import hashlib
import openpyxl
from openpyxl.styles import PatternFill, Font

def calculate_hash(file_path):
    hash_md5 = hashlib.md5()
    with open(file_path, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            hash_md5.update(chunk)
    return hash_md5.hexdigest()

def find_duplicates(directory):
    hash_to_files = {}

    for root, dirs, files in os.walk(directory):
        for filename in files:
            file_path = os.path.join(root, filename)
            file_hash = calculate_hash(file_path)
            if file_hash in hash_to_files:
                hash_to_files[file_hash].append(file_path)
            else:
                hash_to_files[file_hash] = [file_path]

    return {hash_: files for hash_, files in hash_to_files.items() if len(files) > 1}

directory_path = "Y:\Team\DATA" # Provide the desired input folder path
duplicates = find_duplicates(directory_path)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Duplicate Files"

colors = ["76ffff", "ff7676", "ffbb76", "ffff76", "c58aff"]  # Light colors
color_index = 0
pair_id = 1  # Initialize the pair ID

# Set up font for header
header_font = Font(bold=True)

# Write header row with bold font
header_row = ["File Name", "Hash", "Pair ID"]
for col_num, header_text in enumerate(header_row, 1):
    ws.cell(row=1, column=col_num, value=header_text).font = header_font

for hash_, files in duplicates.items():
    fill = PatternFill(start_color=colors[color_index], end_color=colors[color_index], fill_type="solid")
    
    for file_path in files:
        file_name = os.path.basename(file_path)
        ws.cell(row=ws.max_row + 1, column=1, value=file_name).fill = fill
        ws.cell(row=ws.max_row, column=2, value=hash_).fill = fill
        ws.cell(row=ws.max_row, column=3, value=pair_id).fill = fill

    color_index = (color_index + 1) % len(colors)
    pair_id += 1  # Increment the pair ID for the next set of duplicates

excel_filename = "DUP Bank (45S)_1.xlsx"
output_path = "Y:\Team"  # Provide the desired output folder path
output_full_path = os.path.join(output_path, excel_filename)

wb.save(output_full_path)
print(f"Output saved to {output_full_path}")

